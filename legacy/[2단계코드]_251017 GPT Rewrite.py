# translate_with_gpt_after_deepl.py – 2025-09-08 rev-no-qa-tags + all-corpus-monthly + blank-fix
# ------------------------------------------------------------------
# ① …\데이터생성 에서 최신 YYYYMMDD_policysearching.xlsx 자동 로드
# ② GPT-리라이팅 → title_ko / content_ko 2 컬럼 저장  ← (Q&A/Tags 제거)
# ③ Sheet2(“Monthly”)에 content_ko 전체 기반 '최대 10문장' 종합 요약 저장
# ④ Articles 시트(기존 열 + 2개 열) + Monthly 시트 저장 (줄바꿈/열너비 적용)
# ⑤ (추가) 날짜 기준 최근 N일 기사만 별도 시트 저장 + 종합 요약은 해당 시트만 참조
# ------------------------------------------------------------------
import os, glob, time, textwrap, re, calendar
from datetime import datetime
import pandas as pd
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl.styles import Alignment  # Excel 가독성(줄바꿈)
from openpyxl.utils import get_column_letter

# ====== 설정: 최근 N일 ======
RECENT_DAYS = 45  # ← 필요 시 30으로 변경
RECENT_SHEET_NAME = f"Articles_Recent{RECENT_DAYS}d"

# ───────────────────────────────────────── 경로
SRC_DIR = r"C:\Users\cpryul68\OneDrive\전파정책(GPT)\데이터생성"
GPT_DIR = r"C:\Users\cpryul68\OneDrive\전파정책(GPT)\데이터가공(GPT)"
os.makedirs(GPT_DIR, exist_ok=True)

def latest_policy_file(folder: str) -> str | None:
    files = glob.glob(os.path.join(folder, "????????_policysearching.xlsx"))
    return max(files, key=os.path.getmtime) if files else None

INPUT_FILE = latest_policy_file(SRC_DIR)
if not INPUT_FILE:
    raise FileNotFoundError("❌ 최근 *_policysearching.xlsx 파일이 없습니다")

stamp        = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_FILE  = os.path.join(GPT_DIR, f"{stamp}_rewritten.xlsx")

# ───────────────────────────────────────── OpenAI
load_dotenv()
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# 모델 설정(필요 시 여기만 바꾸면 됨)
MODEL_REWRITE = "gpt-4o"       # 기사 리라이팅
MODEL_SUMMARY = "gpt-5"        # 종합 요약(5 사용 불안정 시 자동 폴백)
MODEL_SUMMARY_FALLBACK = "gpt-4.1-mini"  # 빈응답 시 1회 재시도

# ───────────────────────────────────────── Chat Completions 호환 래퍼
_NEW_PARAM_MODELS = re.compile(r"^(gpt-5|o3|o4|gpt-4\.1)")
def chat_create(*, model: str, messages: list, max_tokens: int, temperature: float | None = None):
    kwargs = {"model": model, "messages": messages}
    if _NEW_PARAM_MODELS.match(model):
        kwargs["max_completion_tokens"] = max_tokens
    else:
        kwargs["max_tokens"] = max_tokens
        if temperature is not None:
            kwargs["temperature"] = temperature
    return client.chat.completions.create(**kwargs)

def extract_text(resp) -> str:
    """빈 응답 방지용 안전 추출기."""
    try:
        txt = resp.choices[0].message.content
        if txt is None:
            return ""
        return str(txt)
    except Exception:
        return ""

def clean_text(s: str) -> str:
    # 보이는 공백으로만 구성된 경우 엑셀에서 빈칸처럼 보임 → 정규화
    s = (s or "").replace("\u200b", "").replace("\u2060","").strip()
    # 한 문단 요구 → 줄바꿈은 공백으로 통일
    s = re.sub(r"\s+", " ", s)
    return s

# ====================== (PROMPT: REWRITE – 기존 유지) ======================
SYSTEM_PROMPT = textwrap.dedent("""
You are a senior Korean journalist specializing in radio‑spectrum policy and wireless regulation. Your job is to rewrite English news into Korean that is easy for a middle‑school reader to follow without losing policy accuracy.

OUTPUT LANGUAGE: Korean.

STYLE & STRUCTURE
 • Write the body as one paragraph in this order: Background → Key point → Purpose/implications (6–8 sentences).
 • Use clear, plain Korean (avoid literal translation); keep cause→effect logic; compress redundant phrases.
 • Maintain a neutral newsroom tone (no marketing, no hype, no speculation).

POLICY ORIENTATION
 • Surface: (1) the actor (e.g., FCC, NTIA, Ofcom, ARCEP, BNetzA, ACMA, 일본 총무성(MIC), 과기정통부), (2) the policy instrument (allocation/assignment/auction/consultation/rulemaking), (3) the band(s), (4) the timeline, and (5) the intended policy outcome.
 • Keep distinctions precise: allocation(분배, ITU), allotment(배치), assignment(할당).
 • When mentioned, name band categories (low/mid/mmWave), licensing model (licensed, unlicensed, lightly‑licensed), exclusivity vs. sharing (e.g., LSA, CBRS/SAS/GAA), and coexistence/interference conditions. Do not add facts not in the source.

PRESERVATION RULES
 • Preserve all numbers, units, ranges, and dates exactly as written (e.g., 700 MHz, 3.5 GHz, 24.25–27.5 GHz, 2025-09-08).
 • Keep proper nouns/acronyms; on first use, add the Korean name + English acronym (예: 미국 연방통신위원회(FCC)).
 • Convert relative time terms (“today”, “yesterday”) to explicit dates ONLY if the date is present in the article; otherwise keep the relative term.

FORMATTING (MANDATORY)
 • Output exactly the two blocks below, no markdown, no extra text:
   〈제목〉: concise Korean news headline reflecting the policy action/band/actor.
   〈본문〉: one paragraph, 6–8 sentences.

QUALITY CHECKS
 • Do not invent content, quotes, or numbers.
 • Remove Q&A, tags, disclaimers, or boilerplate if present in the input.
 • Prefer verbs such as “발표했다/검토한다/추진한다/경매를 실시했다/배정했다/협의한다/개정한다”.
 • Additionally, use a standard Korean news-reporting tone (기사보도체): write in concise, objective Korean in a journalistic reporting style, and do not use polite/formal sentence endings such as "~합니다", "~입니다", or "~있습니다".

ADDITIONAL GUARDRAILS FOR SPECTRUM‑POLICY FIDELITY (append to the end of the prompt)

1) Source‑of‑truth scope
• Use only information present in the provided article text. Do not add facts from memory or external sources.
• If a key detail is missing (e.g., band/range, bandwidth, licensing model, timeline, docket/proceeding ID), don't guess.
• When the headline and body conflict, follow the body and state the action status precisely (e.g., “consultation”, “proposal”, “decision/statement”, “auction”, “assignment”).

2) Terminology precision (no conflation)
• Do not conflate allocation (ITU‑level), allotment, and assignment (national licensing).
• Do not infer or rename sharing/licensing frameworks (e.g., CBRS/SAS/GAA, LSA, licence‑exempt) unless explicitly stated in the source.
• Do not normalize bands: preserve frequency ranges exactly as written; do not convert “24.25–27.5 GHz” into “26 GHz band” unless the source uses that label.

3) Numbers & units validation
• Preserve every number, qualifier, and unit verbatim (including “up to”, “at least”, ranges, and decimals).
• Never round, merge, or “sum up” spectrum amounts unless the source provides that figure.
• Keep MHz vs. GHz correct; do not switch units, reorder ranges, or infer contiguity.

4) Actor & jurisdiction accuracy
• Attribute actions to the correct authority; do not swap agencies (e.g., NTIA vs. FCC, Ofcom vs. DCMS).
• If roles differ (e.g., NTIA coordination; FCC rulemaking), reflect each role as stated.
• If the actor is unclear, refer generically to “the authority” rather than inventing a specific body.

5) Forbidden inferences
• Do not assume technology generation (5G/6G), use cases (FWA/NTN), licence type (exclusive/shared/unlicensed), timeline, coverage, or auction revenue unless stated.
• Do not treat consultations as final decisions; keep the status conservative unless the source declares a binding outcome.

6) Output integrity checks (before returning)
• Two blocks only (〈제목〉, 〈본문〉); no extra text.
• Body = 6–8 sentences drawn strictly from the source; if facts are insufficient, prefer the lower bound rather than padding with speculation.
• Ensure all frequency figures, ranges, and key qualifiers present in the source appear in the output; keep docket/proceeding IDs exactly as written (e.g., “WT Docket No. 19‑354”, “FCC 24‑XX”).
• Use the standard Korean news‑reporting tone (기사보도체) with no polite/formal endings (“~합니다/입니다/있습니다” are prohibited).

""").strip()
# ==========================================================================

# ───────────────────────────────────────── GPT 헬퍼 (기사 리라이팅)
def gpt_rewrite(title_en: str, body_en: str) -> str:
    prompt = (
        f"〈Original Title〉\n{title_en.strip()}\n\n"
        f"〈Original Article (English)〉\n{body_en.strip()}\n\n"
        "Rewrite in Korean following the system instructions. Output ONLY the two blocks required."
    )
    rsp = chat_create(
        model=MODEL_REWRITE,
        messages=[{"role":"system","content":SYSTEM_PROMPT},
                  {"role":"user","content":prompt}],
        max_tokens=700,
        temperature=0.2
    )
    return extract_text(rsp).strip()

# ───────────────────────────────────────── 데이터 로드
def load_df(path: str) -> pd.DataFrame:
    xl = pd.ExcelFile(path)
    sheet = "Articles" if "Articles" in xl.sheet_names else xl.sheet_names[0]
    df = pd.read_excel(xl, sheet_name=sheet)
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
    else:
        df["date"] = pd.NaT
    return df

df = load_df(INPUT_FILE)
print(f"📰 {len(df)}개 기사 로딩 – GPT 변환 시작")

# ───────────────────────────────────────── GPT 호출
blocks = []
for i, r in df.iterrows():
    try:
        blk = gpt_rewrite(str(r.get("title", "")), str(r.get("content", "")))
    except Exception as e:
        blk = f"[GPT Error] {e}"
    blocks.append(blk)
    print(f" · {i+1}/{len(df)} 완료")
    time.sleep(1)  # API rate-limit 완충

# ───────────────────────────────────────── 블록 파싱 (Q&A/Tags 제거 버전)
def parse_block(txt: str):
    if txt.startswith("[GPT Error]"):
        return ("[Error]", txt)
    t = textwrap.dedent(txt).strip()
    title = re.search(r"〈제목〉[:：]\s*(.+)", t)
    body  = re.search(r"〈본문〉[:：]\s*(.+)", t, re.S)
    return (
        title.group(1).strip() if title else "",
        re.sub(r"\s+"," ", body.group(1).strip()) if body else ""
    )

parsed = [parse_block(b) for b in blocks]
df[["title_ko","content_ko"]] = pd.DataFrame(parsed, index=df.index)
df.drop(columns=["question_ko","tags"], errors="ignore", inplace=True)

# ───────────────────────────────────────── (신규) 최근 N일 기사 필터링
def filter_recent_articles(df_in: pd.DataFrame, days: int) -> pd.DataFrame:
    """date 열을 기준으로 최근 N일 이내 기사만 반환(날짜 NaT는 제외)."""
    if "date" not in df_in.columns:
        return df_in.iloc[0:0].copy()
    df_work = df_in.copy()
    if not pd.api.types.is_datetime64_any_dtype(df_work["date"]):
        df_work["date"] = pd.to_datetime(df_work["date"], errors="coerce")
    # 오늘(로컬) 기준 N일
    now = pd.Timestamp.now()
    cutoff = now.normalize() - pd.Timedelta(days=days)
    mask = df_work["date"].notna() & (df_work["date"] >= cutoff)
    return df_work.loc[mask].copy()

recent_df = filter_recent_articles(df, RECENT_DAYS)

# ───────────────────────────────────────── 종합 요약 (최근 N일 기사만)
def policy_summary_all(texts_ko: list[str]) -> str:
    corpus = [t for t in texts_ko if isinstance(t, str) and t.strip()]
    joined = "\n\n".join(corpus)[:10000]  # 토큰 안전장치
    if not joined:
        return "데이터가 충분하지 않습니다."
    prompt = textwrap.dedent(f"""
    You are preparing a **monthly spectrum policy digest** for decision‑makers.

    INPUT
    • Below is a concatenated set of Korean article bodies about spectrum/wireless policy (rewritten outputs).
    • Articles may cover multiple countries/regions and regulators.

    GOAL
    • For each country/region detected in the input, **select the 1–2 most consequential items** from a spectrum‑policy perspective, then compress them into **one sentence per country**.
    • Consequential = nationwide allocation/assignment/auction; major consultation/rulemaking; new or expanded sharing frameworks (e.g., CBRS/SAS/GAA, LSA); cross‑border coordination; 6G candidate bands; opening/repurposing ≥100 MHz; actions impacting 6 GHz, 3.x GHz, 24–29 GHz, or comparable strategic bands.
    • If more than 10 countries appear, keep **the top 8–10** by impact/scope and omit the rest.

    COUNTRY/REGULATOR DETECTION (non‑exhaustive examples)
    • U.S. (미국 연방통신위원회(FCC), NTIA), EU/CEPT/ECC, UK(Ofcom), France(ARCEP), Germany(BNetzA), Italy(AGCOM),
      Spain(CNMC), Sweden(PTS), Finland(Traficom), Netherlands(ACM), Canada(ISED), Australia(ACMA),
      Japan(총무성(MIC)), South Korea(과학기술정보통신부(MSIT)), China(MIIT), India(DoT), Singapore(IMDA), etc.
    • Infer country from explicit mentions of regulators/governments/agencies and context within the articles.

    WRITING RULES (OUTPUT LANGUAGE: Korean)
    • Produce **ONE paragraph total**, **≤ 10 sentences**.
    • Write **exactly one sentence per selected country**, merging that country’s 1–2 items via commas/semicolons.
    • Start each sentence with **국가명(주요 규제기관)**, then state: actor → band(s) → policy instrument (allocation/assignment/auction/consultation/rulemaking/sharing) → timeline → intended outcome/implications.
    • Neutral newsroom tone; strictly fact‑based; no speculation or opinions.
    • Preserve all numbers/units/dates/ranges exactly as given (e.g., MHz, GHz, 24.25–27.5 GHz, 2025-09-08).
    • On first mention of a regulator, you may include the Korean full name + English acronym (예: 미국 연방통신위원회(FCC)).
    • Sanitize input: ignore Q&A sections, tags, disclaimers, unrelated device/market commentary.

    IMPORTANT
    • Do **not** invent facts, quotes, or figures.
    • If multiple items exist for one country, **combine them into a single compact sentence**.
    • If content is sparse for a country, include it only if it materially changes spectrum availability or policy direction.

    OUTPUT
    • Return a **single Korean paragraph** (no title or bullets), up to 10 sentences, each sentence covering one country as specified.
    • Additionally, when producing the final summary paragraph, order the sentences by country in the following sequence: South Korea → United States → United Kingdom → Japan → all other countries (in any order after these four).
    
    ARTICLES (Korean, aggregated):
    {joined}
    """)
    rsp = chat_create(
        model=MODEL_SUMMARY,
        messages=[{"role":"system","content":"You are a Korean journalist specializing in radio-spectrum policy and wireless regulation. Output Korean."},
                  {"role":"user","content":prompt}],
        max_tokens=600,
        temperature=None
    )
    raw = extract_text(rsp)
    if not raw or not raw.strip():
        rsp2 = chat_create(
            model=MODEL_SUMMARY_FALLBACK,
            messages=[{"role":"system","content":"You are a Korean journalist specializing in radio-spectrum policy and wireless regulation. Output Korean."},
                      {"role":"user","content":prompt}],
            max_tokens=600,
            temperature=0.2
        )
        raw = extract_text(rsp2)
    cleaned = clean_text(raw)
    return cleaned if cleaned else "요약 생성에 실패했습니다."

# ===== 요약은 '최근 N일' 기사만 참조 =====
recent_texts_ko = recent_df["content_ko"].tolist() if "content_ko" in recent_df.columns else []
monthly_para  = policy_summary_all(recent_texts_ko)

# ───────────────────────────────────────── 저장
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as wr:
    # 1) Articles (전체)
    #   - 전체 시트 저장 전: 날짜를 문자열로 변환
    df_all_out = df.copy()
    if "date" in df_all_out.columns and pd.api.types.is_datetime64_any_dtype(df_all_out["date"]):
        df_all_out["date"] = df_all_out["date"].dt.strftime("%Y-%m-%d")
    df_all_out.to_excel(wr, sheet_name="Articles", index=False)

    # 2) (신규) 최근 N일 시트
    recent_out = recent_df.copy()
    if "date" in recent_out.columns and pd.api.types.is_datetime64_any_dtype(recent_out["date"]):
        recent_out["date"] = recent_out["date"].dt.strftime("%Y-%m-%d")
    recent_out.to_excel(wr, sheet_name=RECENT_SHEET_NAME, index=False)

    # 3) Monthly (요약 = 최근 N일 기사만 기반)
    month_name = calendar.month_name[int(datetime.now().strftime('%m'))]
    title = f"{datetime.now().year}년 {datetime.now().month}월 최근 주파수 정책 종합 요약"
    monthly_df = pd.DataFrame({"Monthly Summary":[title], "Content":[monthly_para]})
    monthly_df.to_excel(wr, sheet_name="Monthly", index=False)

    # 보기 좋게: Monthly 시트 줄바꿈/열너비 적용
    ws = wr.sheets["Monthly"]
    ws.column_dimensions[get_column_letter(1)].width = 36   # Monthly Summary
    ws.column_dimensions[get_column_letter(2)].width = 120  # Content
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

print(f"\n✅ 변환 완료 → {OUTPUT_FILE}\n"
      f"   · 전체 기사 시트: Articles\n"
      f"   · 최근 {RECENT_DAYS}일 시트: {RECENT_SHEET_NAME}\n"
      f"   · Monthly 요약은 최근 {RECENT_DAYS}일 기사만 반영")
